# geocode_broker_excel.py
# Read:  C:\ngrok\RB_DataMining\rawdata\broker_dimensions_master_V2.xlsx (sheet: broker_master_enriched)
# Input: column F = address
# Output: write Latitude to H, Longitude to I
# Save as: broker_dimensions_master_V3.xlsx (same folder unless you change OUT_PATH)

import os
import json
import time
import argparse
from pathlib import Path
from typing import Dict, Any, Optional, Tuple

import requests
from openpyxl import load_workbook


GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"


def geocode_google(address: str, api_key: str, session: requests.Session, region: str = "tw", language: str = "zh-TW") -> Dict[str, Any]:
    params = {
        "address": address,
        "key": api_key,
        "region": region,
        "language": language,
    }
    r = session.get(GEOCODE_URL, params=params, timeout=20)
    r.raise_for_status()
    return r.json()


def extract_lat_lng(resp: Dict[str, Any]) -> Tuple[Optional[float], Optional[float], str]:
    status = resp.get("status", "ERROR")
    if status == "OK" and resp.get("results"):
        loc = resp["results"][0]["geometry"]["location"]
        return float(loc["lat"]), float(loc["lng"]), status
    return None, None, status


def load_cache(cache_path: Path) -> Dict[str, Dict[str, Any]]:
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache_path: Path, cache: Dict[str, Dict[str, Any]]) -> None:
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_addr(s: str) -> str:
    return " ".join((s or "").strip().split())


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="in_path",
                        default=r"C:\ngrok\RB_DataMining\rawdata\broker_dimensions_master_V2.xlsx")
    parser.add_argument("--out", dest="out_path", default="")
    parser.add_argument("--sheet", default="broker_master_enriched")
    parser.add_argument("--api_key", default="", help="Google Geocoding API key (or set env GOOGLE_MAPS_API_KEY)")
    parser.add_argument("--sleep", type=float, default=0.12, help="Seconds to sleep between requests (avoid rate issues)")
    parser.add_argument("--max_retries", type=int, default=5)
    parser.add_argument("--retry_backoff", type=float, default=1.6)
    args = parser.parse_args()

    api_key = args.api_key or os.getenv("GOOGLE_MAPS_API_KEY") or "AIzaSyCSZEA-ruY_Ks5iruFShgPZBbABrrp864o"

    if not api_key:
        raise SystemExit("Missing API key. Provide --api_key or set GOOGLE_MAPS_API_KEY environment variable.")

    in_path = Path(args.in_path)
    if not in_path.exists():
        raise SystemExit(f"Input file not found: {in_path}")

    out_path = Path(args.out_path) if args.out_path else in_path.with_name("broker_dimensions_master_V3.xlsx")

    cache_path = out_path.with_suffix(".geocode_cache.json")
    cache = load_cache(cache_path)

    wb = load_workbook(in_path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}. Available: {wb.sheetnames}")
    ws = wb[args.sheet]

    # Columns: F=6 address, H=8 lat, I=9 lng
    COL_ADDR = 6
    COL_LAT = 8
    COL_LNG = 9

    session = requests.Session()

    # Iterate rows: assumes header in row 1. If not, change start_row=1.
    start_row = 2
    max_row = ws.max_row

    updated = 0
    skipped = 0
    failed = 0

    for r in range(start_row, max_row + 1):
        addr_cell = ws.cell(row=r, column=COL_ADDR)
        raw_addr = addr_cell.value

        if raw_addr is None or str(raw_addr).strip() == "":
            skipped += 1
            continue

        addr = normalize_addr(str(raw_addr))

        lat_cell = ws.cell(row=r, column=COL_LAT)
        lng_cell = ws.cell(row=r, column=COL_LNG)

        # If already has both lat/lng, skip
        if lat_cell.value not in (None, "") and lng_cell.value not in (None, ""):
            skipped += 1
            continue

        # Cache hit
        if addr in cache and cache[addr].get("status") == "OK":
            lat_cell.value = cache[addr]["lat"]
            lng_cell.value = cache[addr]["lng"]
            updated += 1
            continue

        # Call API with retries
        attempt = 0
        while True:
            attempt += 1
            try:
                resp = geocode_google(addr, api_key, session=session)
                lat, lng, status = extract_lat_lng(resp)

                cache[addr] = {
                    "status": status,
                    "lat": lat,
                    "lng": lng,
                    "timestamp": time.time(),
                }

                if status == "OK" and lat is not None and lng is not None:
                    lat_cell.value = lat
                    lng_cell.value = lng
                    updated += 1
                else:
                    # keep blank for review
                    failed += 1

                break

            except requests.HTTPError as e:
                # Handle quota / transient
                if attempt >= args.max_retries:
                    cache[addr] = {"status": "HTTP_ERROR", "error": str(e), "timestamp": time.time()}
                    failed += 1
                    break
                time.sleep(args.retry_backoff ** attempt)

            except Exception as e:
                if attempt >= args.max_retries:
                    cache[addr] = {"status": "ERROR", "error": str(e), "timestamp": time.time()}
                    failed += 1
                    break
                time.sleep(args.retry_backoff ** attempt)

        # Save cache periodically
        if (updated + failed) % 50 == 0:
            save_cache(cache_path, cache)

        time.sleep(args.sleep)

    # Final save
    save_cache(cache_path, cache)
    wb.save(out_path)

    print(f"Done. Saved: {out_path}")
    print(f"Updated: {updated}, Skipped: {skipped}, Failed: {failed}")
    print(f"Cache: {cache_path}")


if __name__ == "__main__":
    main()
